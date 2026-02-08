from openpyxl import load_workbook
from django.db import transaction
from .models import Product, Supplier, Category


def _get_or_create_supplier(name):
    if not name:
        return None
    supplier, _ = Supplier.objects.get_or_create(name=str(name).strip())
    return supplier


def _get_or_create_category(name):
    if not name:
        return None
    category, _ = Category.objects.get_or_create(name=str(name).strip())
    return category


@transaction.atomic
def import_products_from_excel(file_obj, mark_discontinued=False):
    """
    Excelファイルを読み込み、商品データを登録 or 廃盤フラグを立てる。
    """
    wb = load_workbook(filename=file_obj, data_only=True)
    ws = wb.active

    created, updated, skipped = 0, 0, 0
    row = 4  # B4～E11が対象範囲という仕様

    while True:
        code = ws[f'B{row}'].value
        name = ws[f'C{row}'].value
        supplier_name = ws[f'D{row}'].value
        category_name = ws[f'E{row}'].value

        if not code:
            break  # 商品コードが空なら終了

        code = str(code).strip()

        if mark_discontinued:
            try:
                p = Product.objects.get(product_code=code)
                if not p.discontinued:
                    p.discontinued = True
                    p.save()
                    updated += 1
                else:
                    skipped += 1
            except Product.DoesNotExist:
                skipped += 1
        else:
            supplier = _get_or_create_supplier(supplier_name)
            category = _get_or_create_category(category_name)
            p, created_flag = Product.objects.update_or_create(
                product_code=code,
                defaults={
                    'name': name or '',
                    'supplier': supplier,
                    'category': category,
                    'discontinued': False,
                }
            )
            if created_flag:
                created += 1
            else:
                updated += 1
        row += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "mark_discontinued": mark_discontinued,
    }
